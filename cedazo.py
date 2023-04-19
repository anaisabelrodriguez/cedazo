#!/bin/env python
# Cedazo - Morph, Meld and Merge data - Copyright © 2023 Iwan van der Kleijn - See LICENSE.txt for conditions
import openpyxl

from data import cambiar_coma, change_colour, change_value, format_column, format_column_iter, format_condition, format_condition_iter, format_condition_iter2, insert_column, quitar_acentos, remove, unmerge_cells
from miargparse import parser

from specific import assign_value, comprobar_filas, remove_draft 
#Damos la localización del fichero de entrada
ruta_input = "C:\\temp\\in.xlsx"
#Damos la localización del fichero de salida
ruta_output = "C:\\temp\\out.xlsx"
# Creamos objeto wb (libro) de tipo workbook y lo cargamos con lo del excel
#args = parser.parse_args()
#ruta_input = args.ruta_input
#print(args)

wb = openpyxl.load_workbook(ruta_input)
# Creamos objeto ws (hoja), siendo la hoja activa
ws = wb.active

#Metodo que desmergea las celdas de las filas a eliminar
unmerge_cells(ws)

# Metodo que sirve para borrar todas las filas de 1 a 11
remove(ws,1,11)

# Metodo que elimina el color amarillo de todas la celdas que sean amarillas 
change_colour(ws)

#Método que elimina las filas en cuya columna C(TeamRequestStatus) se encuentre a 'Draft'
remove_draft(ws)

#Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera y copiando el formato de la columna H
insert_column(ws, colNr=9, headerRow=1, headerVal='FTES. Pdtes.')

#Metodo que da formato a la columna que se ha creado usando range
format_column(ws, colNr= 9)
#Metodo que da formato a la columna que se ha creado utilizando iter_cols
#format_column_iter(ws, colNr= 9)

#Se cambia la cabecera de la columna F de 'Additional Notes' a 'CRITICIDAD'
change_value(ws,1,6,"CRITICIDAD",None,0)

#Se cambia la cabecera de la columna O(del fichero original) de 'Team Request Comment 1' a 'CLIENTE'
change_value(ws,1,15,"CLIENTE",None,0)

#Se cambian las comas decimales por comas para que realice bien las operaciones.
cambiar_coma(ws,7)
cambiar_coma(ws,8)

#Se asigna valor a la nueva columna que nos hemos creado
assign_value(ws)

#Se quitan los acentos de la columna que vamos a tratar.
quitar_acentos(ws,6)

comprobar_filas(ws)

#Metodo que da formato a la columna que se ha creado según criterio  
#color_yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type = "solid")
#format_condition(ws, colNr= 9, condition="CRITICA", color=color_yellow)
#Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y enumerate
#format_condition_iter(ws, colNr= 9, condition="CRITICA", color=color_yellow)
#Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y cell
#format_condition_iter2(ws, colNr= 9, condition="CRITICA", color=color_yellow)

# Save the workbook to the output file
wb.save(ruta_output)


