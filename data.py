from copy import copy
from types import CellType
import openpyxl
from openpyxl.styles import PatternFill 


def remove_empty(ws): 
    '''Metodo que sirve para borrar todas las filas que estan vacias (no se utiliza de momento porque no sirve para borrar
       lad filas que hay que eliminar pero tienen contenido)'''
    filas = ws.max_row
    for i in range(filas, 0, -1):
        celdas_vacias = all([cell.value is None for cell in ws[i]])
        if celdas_vacias:
            ws.delete_rows(i, 1)

def unmerge_cells(ws):
    '''Metodo que desmergea las celdas de las filas a eliminar'''
    # Buscamos las celdas que estan mergeadas
    for merge in list(ws.merged_cells):
        # Separamos esas celdas mergeadas
        ws.unmerge_cells(range_string=str(merge))

def remove(ws,ini,end): 
    '''Metodo que sirve para borrar todas las filas de 1 a 11'''
    ws.delete_rows(ini,end)


def change_colour(ws): 
    '''Metodo que a todas la celdas les pone color blanco (quita el amarillo) '''
    max_row = ws.max_row
    #print('El numero de filas rellenas es: ', max_row)
    max_column = ws.max_column
    #print('El numero de columnas rellenas es: ', max_column)

    # Hacemos doble bucle para recoger todos los datos de la tabla que empezaba en fila 12 (ahora empezara en la tabla 
    # destino en la fila 1).
    for i_row in range(1, max_row + 1):
       for i_column in range(1, max_column + 1):
           cell = ws.cell(row = i_row, column = i_column)
           #If ((cell.fill.fgcolor.type == 'indexed' and cell.fill.fgcolor.indexed == 43) or
           #(cell.fill.fgcolor.type == 'rgb' and cell.fill.fgcolor.rgb == 'FFFFFF99')):
           # Cambia para todas las celdas el fondo a blanco
           cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type = "solid")

def copyStyle(newCell, cell): 
    '''Metodo que copia el formato de una celda a una nueva '''
    if cell.has_style: 
        newCell.style = copy(cell.style) 
        newCell.font = copy(cell.font) 
        newCell.border = copy(cell.border) 
        newCell.fill = copy(cell.fill) 
        newCell.number_format = copy(cell.number_format) 
        newCell.protection = copy(cell.protection) 
        newCell.alignment = copy(cell.alignment)
  
def insert_column(ws, colNr ,headerRow , headerVal):
    '''Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera'''
    ws.insert_cols(colNr)
    #añadimos el título a la columna 
    ws.cell(row=headerRow, column=colNr).value = headerVal

def format_column(ws, colNr):
    '''Metodo que da formato a la columna que se ha creado usando método range'''
    max_row = ws.max_row
    #recorremos todas las celdas de la columna H para copiar su formato a la nueva columna mediante el método
    #copyStyle definido más arriba
    for i_row in range(1, max_row + 1):
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_origin = ws.cell(row=i_row, column=colNr-1)
        copyStyle(cell_new, cell_origin)

def format_column_iter(ws, colNr):  
    '''Metodo que da formato a la columna que se ha creado usando método iter_cols'''
    #iter_cols devuelve un generador de tuplas, donde cada tupla contiene todas las celdas de una columna en particular, 
    #desde la fila inicial hasta la fila final (establecidas con los argumentos min_row y max_row, respectivamente).
    #itera a través de todas las columnas de la hoja de trabajo (ws) y 
    #selecciona solo la columna anterior a la que se desea copiar (colNr-1).
    for col in ws.iter_cols(min_row=1, min_col=colNr-1, max_col=colNr-1): 
        #recorre todas las celdas de esa columna y copia su estilo a la celda 
        #correspondiente en la nueva columna (colNr).
        #recorre todas las celdas de la columna seleccionada en la variable col, y en cada iteración, 
        #la variable cell_origin se asigna a una celda de la columna actual. 
        for cell_origin in col:
            print("la fila a tratar es: ", cell_origin.row)
            print("la coordenada a tratar es: ", cell_origin.coordinate)
            cell_new = ws.cell(row=cell_origin.row, column=colNr)
            copyStyle(cell_new, cell_origin)
        

def format_condition(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio'''   
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
     #columna nueva (I - FTES. Pdtes.) a amarillo
    max_row = ws.max_row    
    for i_row in range(1, max_row + 1):   
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_condition = ws.cell(row=i_row, column= colNr-3)
        if cell_condition.value == condition:
           cell_new.fill = color

def format_condition_iter(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y enumerate'''  
    #colNr: es el número de la columna que se desea formatear.
    #condition: es el valor que se debe buscar en la columna colNr-3 para aplicar el formato condicional.
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
    #columna nueva (I - FTES. Pdtes.) a amarillo
    #values_only=True ->indica si solo se deben recorrer los valores de las celdas sin objetos "Cell"
    #si utilizamos values_only=True necesitamos el enumerate, si no lo utilizamos luego tendremos que acceder
    #al atributo .value
    for col in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3, values_only=True): 
        print(col)
        #col tiene ('Additional Notes', 'CRITICA', 'URGENTE', 'URGENTE', None, 'CRITICA', ',')
        #el siguiente for itera a través de cada celda en la columna seleccionada
        #La función enumerate() devuelve una tupla con un índice i y un valor value correspondiente al 
        #valor de la celda en la posición i.
        for i, value in enumerate(col):
        #Si el valor de la celda coincide con la condición (value == condition), entonces el código cambia el colorç
        #de fondo de la celda correspondiente en la columna colNr usando el método fill de la celda.   
           if value == condition:
              print("el valor de lo encontrado es:", value)
              print("el valor de i es:", i)
              #en cell_coor almacenamos la coordenada de la celda en la columna correspondiente (colNr) y 
              #la fila correspondiente (i + 1). Esta coordenada se almacena en la variable cell_coor.
              #utilizamos i+1 porque aquí empezamos desde 0 y en excel se empieza desde 1
              cell_coor = ws.cell(row=i+1,column= colNr).coordinate
              print('coordenada: ', cell_coor)
              #Se utiliza la coordenada de la celda (cell_coor) para obtener un objeto Cell de la hoja de cálculo (cell) 
              #correspondiente a la celda en la fila y columna especificadas.
              cell = ws[cell_coor]
              cell.fill = color      

def format_condition_iter2(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y cell'''  
    #colNr: es el número de la columna que se desea formatear.
    #condition: es el valor que se debe buscar en la columna colNr-3 para aplicar el formato condicional.
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
    #columna nueva (I - FTES. Pdtes.) a amarillo
    #si no recuperamos el valor con values_only=True tendremos que acceder luego al 
    #al atributo .value
    for col_condition in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3): 
        print(col_condition)
        #col tiene ('Additional Notes', 'CRITICA', 'URGENTE', 'URGENTE', None, 'CRITICA', ',')
        #el siguiente for itera a través de cada celda en la columna seleccionada
        for cell_condition in col_condition:
        #Si el valor de la celda coincide con la condición (value == condition), entonces el código cambia el colorç
        #de fondo de la celda correspondiente en la columna colNr usando el método fill de la celda.   
           if cell_condition.value == condition:
              print('entra en if del condition con coordinada de celda:', cell_condition.coordinate)
              print('La fila de la celda de la condicion es: ',cell_condition.row)
              print("el valor de lo encontrado es:", cell_condition.value)
              #en cell_coor almacenamos la coordenada de la celda 
              # construyo la coordenada de la celda a la que quiero cambiar el formato, a partir de la fila
              # (misma fila que la celda de la condicion) y la columna a cambiar formato (pasada por parametro)
              cell_coor = ws.cell(row=cell_condition.row, column=colNr).coordinate
              print('coordenada: ', cell_coor)
              #Se utiliza la coordenada de la celda (cell_coor) para obtener un objeto cell de la hoja de cálculo (ws) 
              #correspondiente a la celda en la fila y columna especificadas.
              cell = ws[cell_coor]
              #se cambia el color al nuevo objeto creado
              cell.fill = color                    